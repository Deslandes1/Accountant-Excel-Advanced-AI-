import streamlit as st
import pandas as pd
import sqlite3
import datetime
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils.dataframe import dataframe_to_rows

# ====== CHECK FOR EDGE TTS (REQUIRED FOR NATIVE VOICES) ======
EDGE_TTS_AVAILABLE = False
try:
    import edge_tts
    import asyncio
    import tempfile
    import os
    EDGE_TTS_AVAILABLE = True
except ImportError:
    pass

# ----------------------------------------------------------------------
# Page config
# ----------------------------------------------------------------------
st.set_page_config(page_title="Excel Advanced Accounting", layout="wide")

# ----------------------------------------------------------------------
# Custom CSS – Blue Theme + Full Table Styling
# ----------------------------------------------------------------------
st.markdown("""
<style>
    .stApp { background-color: #e6f2ff !important; }
    .stApp [data-testid="stAppViewContainer"] { background-color: transparent !important; }
    [data-testid="stSidebar"] {
        background-color: #cce5ff !important;
        border-right: 1px solid #99ccff;
    }
    [data-testid="stSidebar"] * { color: #003366 !important; }
    h1, h2, h3 { color: #003366 !important; }
    .stTextInput > div > div > input,
    .stTextArea > div > textarea,
    .stSelectbox > div > div {
        background: #ffffff !important;
        color: #003366 !important;
        border: 1px solid #99ccff !important;
        border-radius: 8px !important;
    }
    .stButton > button {
        background: linear-gradient(105deg, #1e88e5 0%, #42a5f5 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 40px !important;
        padding: 0.6rem 2rem !important;
        font-weight: 600 !important;
        width: 100% !important;
        transition: 0.2s;
    }
    .stButton > button:hover {
        transform: scale(1.02);
        box-shadow: 0 4px 20px rgba(30, 136, 229, 0.4);
    }
    [data-testid="stMetricValue"] { color: #003366 !important; }
    .stTabs [data-baseweb="tab-list"] {
        background-color: #cce5ff !important;
        border-radius: 8px !important;
    }
    .stTabs [data-baseweb="tab"] { color: #003366 !important; }
    .stTabs [aria-selected="true"] {
        background-color: #1e88e5 !important;
        color: white !important;
        border-radius: 8px !important;
    }
    div[data-testid="stDataFrame"] {
        border: 1px solid #b0c4de !important;
        border-radius: 4px !important;
        overflow: hidden !important;
    }
    div[data-testid="stDataFrame"] table {
        border-collapse: collapse !important;
        width: 100% !important;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif !important;
        font-size: 14px !important;
        border: 1px solid #b0c4de !important;
        background-color: #ffffff !important;
    }
    div[data-testid="stDataFrame"] thead tr th {
        background-color: #1e88e5 !important;
        color: white !important;
        font-weight: bold !important;
        text-align: center !important;
        padding: 8px 6px !important;
        border: 1px solid #1565c0 !important;
    }
    div[data-testid="stDataFrame"] tbody tr:nth-child(even) {
        background-color: #f0f8ff !important;
    }
    div[data-testid="stDataFrame"] tbody tr:nth-child(odd) {
        background-color: #ffffff !important;
    }
    div[data-testid="stDataFrame"] tbody tr:hover {
        background-color: #d9eaf7 !important;
    }
    div[data-testid="stDataFrame"] td {
        padding: 6px 8px !important;
        border: 1px solid #b0c4de !important;
        text-align: right !important;
        color: #1a2a3a !important;
    }
    div[data-testid="stDataFrame"] td:first-child {
        text-align: left !important;
    }
    div[data-testid="stDataFrame"] td:nth-child(3) {
        text-align: left !important;
    }
</style>
""", unsafe_allow_html=True)

# ----------------------------------------------------------------------
# Translations (for other tabs – Reconciliation uses hardcoded English)
# ----------------------------------------------------------------------
translations = {
    "en": {
        "app_title": "Excel Advanced Accounting",
        "subtitle": "Professional Accounting & Loan Management Suite",
        "login_title": "🔐 Login",
        "login_password": "Enter password to unlock",
        "wrong_password": "Wrong password. Access denied.",
        "logout": "🚪 Logout",
        "dashboard": "📊 Dashboard",
        "cash_tab": "💰 Cash In/Out",
        "loans_tab": "🏦 Loans",
        "reports_tab": "📄 Reports",
        "reconciliation_tab": "📋 Reconciliation Ledger",
        "current_balance": "Current Cash Balance",
        "current_balance_htg": "Current Cash Balance (HTG)",
        "recent_transactions": "Recent Cash Transactions",
        "active_loans": "Active Loans",
        "no_active_loans": "No active loans.",
        "add_transaction": "Add Transaction",
        "date": "Date",
        "type": "Type",
        "income": "Income",
        "expense": "Expense",
        "category": "Category (e.g., Sales, Rent, Salary)",
        "description": "Description",
        "amount": "Amount ($)",
        "amount_htg": "Amount (HTG)",
        "transaction_added": "Transaction added!",
        "transaction_history": "Transaction History",
        "download_excel": "📥 Download Excel",
        "loan_management": "Loan Management",
        "add_new_loan": "➕ Add New Loan",
        "borrower_name": "Borrower Name",
        "loan_amount": "Loan Amount ($)",
        "loan_amount_htg": "Loan Amount (HTG)",
        "start_date": "Start Date",
        "interest_rate": "Interest Rate (%)",
        "payment_frequency": "Payment Frequency",
        "weekly": "Weekly",
        "monthly": "Monthly",
        "payment_amount": "Payment Amount ($)",
        "payment_amount_htg": "Payment Amount (HTG)",
        "total_payments": "Total Number of Payments",
        "create_loan": "Create Loan",
        "loan_created": "Loan created!",
        "all_loans": "All Loans",
        "select_loan": "Select Loan ID to record payment or view details",
        "remaining_payments": "Remaining payments",
        "status": "Status",
        "record_payment": "Record Payment",
        "payment_date": "Payment Date",
        "payment_recorded": "Payment recorded!",
        "payment_history": "Payment History",
        "no_loans": "No loans yet. Add a loan above.",
        "generate_reports": "Generate Professional Reports",
        "report_type": "Report Type",
        "cash_flow_statement": "Cash Flow Statement",
        "loan_status_report": "Loan Status Report",
        "payment_history_report": "Payment History Report",
        "generate": "Generate",
        "from_date": "Start Date",
        "to_date": "End Date",
        "total_income": "Total Income",
        "total_expense": "Total Expense",
        "net_cash_flow": "Net Cash Flow",
        "filter_by_status": "Filter by status",
        "all": "All",
        "active": "active",
        "completed": "completed",
        "no_data": "No data available.",
        "select_loan_for_history": "Select Loan",
        "created_by": "Python Developer",
        "reconciliation_title": "Reconciliation July - 2026",
        "exchange_rate": "Exchange Rate: 1 USD = 100 HTG",
        "balance_usd": "Balance USD",
        "balance_htg": "Balance HTG",
        "credit_cash_in": "Credit (Cash In HTG)",
        "credit_cash_in_usd": "Credit (Cash In USD)",
        "description_item": "Description / Item Details",
        "qty": "qty",
        "currency_htg": "Currency Unit (HTG)",
        "unit_htg": "unit htg",
        "unit_usd": "unit usd",
        "total_htg": "total htg",
        "total_usd": "total usd",
        "add_entry": "Add Entry",
        "credit": "Credit (Cash In HTG)",
        "qty_input": "Quantity",
        "unit_htg_input": "Unit Price (HTG)",
        "description_input": "Description / Item Details",
        "entry_added": "Entry added!",
        "download_reconciliation": "📥 Download Excel",
        "starting_balance_usd": "Starting Balance (USD)",
        "starting_balance_htg": "Starting Balance (HTG)",
        "initial_balance_forwarded": "Balance Forwarded from February",
        "delete_entry": "Delete Entry",
        "cannot_delete_balance": "Cannot delete the initial balance row.",
        "voice_welcome": "Welcome to Excel Advanced Accounting.",
        "voice_ledger": "Here is your Reconciliation Ledger summary.",
        "voice_entries": "Currently, you have {count} entries.",
        "voice_credit": "Total cash in is {credit_htg:,.2f} HTG, which is {credit_usd:,.2f} USD.",
        "voice_expenses": "Total expenses are {expense_htg:,.2f} HTG and {expense_usd:,.2f} USD.",
        "voice_balance": "Your current net balance is {balance_htg:,.2f} HTG and {balance_usd:,.2f} USD. This is calculated as total cash in minus total expenses.",
        "voice_how_it_works": "Remember: each cash‑in increases your net balance, and each purchase decreases it. The system automatically converts HTG to USD using the exchange rate of 1 USD = 100 HTG.",
        "voice_closing": "You can download a professionally formatted Excel report with one click. This application was built by Gesner Deslandes, Chief Engineer at GlobalInternet.py."
    },
    "fr": {
        # ... (keep your existing French translations; ensure all keys used in other tabs exist)
    },
    "es": {
        # ... (keep your existing Spanish translations)
    }
}

def _(key):
    lang = st.session_state.get("language", "en")
    return translations[lang].get(key, key)

# ----------------------------------------------------------------------
# Authentication
# ----------------------------------------------------------------------
def get_expected_password():
    try:
        return st.secrets["password"]
    except KeyError:
        return "20082010"

def check_password():
    def password_entered():
        if st.session_state["password"] == get_expected_password():
            st.session_state["authenticated"] = True
            del st.session_state["password"]
        else:
            st.session_state["authenticated"] = False

    if "authenticated" not in st.session_state:
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            st.image(
                "https://raw.githubusercontent.com/Deslandes1/Accountant-Excel-Advanced-AI-/main/Gemini_Generated_Image_8s108y8s108y8s10.png",
                width=100
            )
        with col2:
            st.markdown(f"<h1 style='text-align: center;'>{_('app_title')}</h1>", unsafe_allow_html=True)
            st.markdown(f"<p style='text-align: center;'><em>{_('subtitle')}</em></p>", unsafe_allow_html=True)
        with col3:
            st.markdown("""
            <div style='text-align: right;'>
                <b>GlobalInternet.py</b><br>
                Gesner Deslandes<br>
                Chief Engineer at GlobalInternet.py
            </div>
            """, unsafe_allow_html=True)
        st.divider()
        st.text_input(_("login_password"), type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["authenticated"]:
        st.text_input(_("login_password"), type="password", on_change=password_entered, key="password")
        st.error(_("wrong_password"))
        return False
    else:
        return True

def logout():
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

# ----------------------------------------------------------------------
# Database setup
# ----------------------------------------------------------------------
def init_db():
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS cash_transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        type TEXT,
        category TEXT,
        description TEXT,
        amount REAL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS loans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        borrower TEXT,
        amount REAL,
        start_date TEXT,
        interest_rate REAL,
        payment_frequency TEXT,
        payment_amount REAL,
        total_payments INTEGER,
        payments_made INTEGER DEFAULT 0,
        status TEXT DEFAULT 'active'
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS loan_payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        loan_id INTEGER,
        payment_date TEXT,
        amount REAL,
        FOREIGN KEY (loan_id) REFERENCES loans (id)
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS reconciliation_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        credit REAL DEFAULT 0,
        description TEXT,
        qty REAL DEFAULT 0,
        unit_htg REAL DEFAULT 0,
        unit_usd REAL DEFAULT 0,
        total_htg REAL DEFAULT 0,
        total_usd REAL DEFAULT 0
    )""")
    c.execute("SELECT COUNT(*) FROM reconciliation_entries")
    if c.fetchone()[0] == 0:
        demo_entries = [
            ("2023-03-01", "Cash in from operation (HTG)", 300000.00, 0, 0, 0, 0, 0),
            ("2023-03-02", "Office supplies - paper & pens", 0, 20, 150.00, 1.50, 3000.00, 30.00),
            ("2023-03-03", "Equipment rental - projector", 0, 5, 500.00, 5.00, 2500.00, 25.00),
            ("2023-03-05", "Fuel for delivery vehicles", 0, 40, 125.00, 1.25, 5000.00, 50.00),
            ("2023-03-07", "Cash in from sales (HTG)", 120000.00, 0, 0, 0, 0, 0),
            ("2023-03-10", "Utility bills - electricity", 0, 0, 0, 0, 4000.00, 40.00),
            ("2023-03-12", "Transportation - maintenance", 0, 2, 800.00, 8.00, 1600.00, 16.00)
        ]
        for entry in demo_entries:
            c.execute("""INSERT INTO reconciliation_entries (date, description, credit, qty, unit_htg, unit_usd, total_htg, total_usd)
                         VALUES (?, ?, ?, ?, ?, ?, ?, ?)""", entry)
        conn.commit()
    conn.commit()
    conn.close()

init_db()

# ----------------------------------------------------------------------
# Helper functions
# ----------------------------------------------------------------------
def add_cash_transaction(date, trans_type, category, description, amount):
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("INSERT INTO cash_transactions (date, type, category, description, amount) VALUES (?,?,?,?,?)",
              (date, trans_type, category, description, amount))
    conn.commit()
    conn.close()

def get_cash_balance():
    conn = sqlite3.connect("accounting.db")
    df = pd.read_sql_query("SELECT type, amount FROM cash_transactions", conn)
    conn.close()
    if df.empty:
        return 0
    income = df[df['type'] == 'Income']['amount'].sum()
    expense = df[df['type'] == 'Expense']['amount'].sum()
    return income - expense

def get_cash_flow(start_date, end_date):
    conn = sqlite3.connect("accounting.db")
    df = pd.read_sql_query("SELECT * FROM cash_transactions WHERE date BETWEEN ? AND ?", conn, params=(start_date, end_date))
    conn.close()
    return df

def add_loan(borrower, amount, start_date, interest_rate, payment_frequency, payment_amount, total_payments):
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("""INSERT INTO loans (borrower, amount, start_date, interest_rate, payment_frequency, payment_amount, total_payments)
                 VALUES (?,?,?,?,?,?,?)""",
              (borrower, amount, start_date, interest_rate, payment_frequency, payment_amount, total_payments))
    conn.commit()
    conn.close()

def record_loan_payment(loan_id, payment_date, amount):
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("INSERT INTO loan_payments (loan_id, payment_date, amount) VALUES (?,?,?)", (loan_id, payment_date, amount))
    c.execute("UPDATE loans SET payments_made = payments_made + 1 WHERE id = ?", (loan_id,))
    c.execute("SELECT payments_made, total_payments FROM loans WHERE id = ?", (loan_id,))
    made, total = c.fetchone()
    if made >= total:
        c.execute("UPDATE loans SET status = 'completed' WHERE id = ?", (loan_id,))
    conn.commit()
    conn.close()

def get_loans(status=None):
    conn = sqlite3.connect("accounting.db")
    query = "SELECT * FROM loans"
    if status:
        query += " WHERE status = ?"
        df = pd.read_sql_query(query, conn, params=(status,))
    else:
        df = pd.read_sql_query(query, conn)
    conn.close()
    return df

def get_loan_payments(loan_id):
    conn = sqlite3.connect("accounting.db")
    df = pd.read_sql_query("SELECT * FROM loan_payments WHERE loan_id = ? ORDER BY payment_date", conn, params=(loan_id,))
    conn.close()
    return df

def get_reconciliation_entries():
    conn = sqlite3.connect("accounting.db")
    df = pd.read_sql_query(
        "SELECT id, date, credit, description, qty, unit_htg, unit_usd, total_htg, total_usd FROM reconciliation_entries ORDER BY id",
        conn)
    conn.close()
    if not df.empty:
        # Calculate net balance (cumulative credit - cumulative expenses)
        net_htg = []
        net_usd = []
        running_net_htg = 0
        running_net_usd = 0
        for idx, row in df.iterrows():
            running_net_htg += row['credit'] - row['total_htg']
            running_net_usd += (row['credit'] / 100) - row['total_usd']
            net_htg.append(running_net_htg)
            net_usd.append(running_net_usd)
        df['net_htg'] = net_htg
        df['net_usd'] = net_usd
    else:
        df = pd.DataFrame(columns=['id', 'date', 'credit', 'description', 'qty', 'unit_htg', 'unit_usd',
                                   'total_htg', 'total_usd', 'net_htg', 'net_usd'])
    return df

def add_reconciliation_entry(date, credit_htg, description, qty, unit_htg, unit_usd, total_htg, total_usd):
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("""INSERT INTO reconciliation_entries (date, credit, description, qty, unit_htg, unit_usd, total_htg, total_usd)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
              (date, credit_htg, description, qty, unit_htg, unit_usd, total_htg, total_usd))
    conn.commit()
    conn.close()

def delete_reconciliation_entry(entry_id):
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("DELETE FROM reconciliation_entries WHERE id = ?", (entry_id,))
    conn.commit()
    conn.close()

def reset_ledger():
    conn = sqlite3.connect("accounting.db")
    c = conn.cursor()
    c.execute("DELETE FROM reconciliation_entries")
    conn.commit()
    conn.close()
    st.cache_data.clear()

def generate_pdf_report(title, data, columns):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 12))
    if not data.empty:
        table_data = [columns] + data.values.tolist()
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph(_("no_data"), styles['Normal']))
    doc.build(story)
    buffer.seek(0)
    return buffer

def usd_to_htg(usd):
    return usd * 100

EXCHANGE_RATE = 100

# ----------------------------------------------------------------------
# AI Voice Functions
# ----------------------------------------------------------------------
def generate_voice_explanation(entries, net_htg, net_usd, lang='en'):
    if entries.empty:
        return "There are no entries in the ledger. Please add a transaction."
    
    total_credit_htg = entries['credit'].sum()
    total_expense_htg = entries['total_htg'].sum()
    total_expense_usd = entries['total_usd'].sum()
    total_credit_usd = total_credit_htg / EXCHANGE_RATE
    
    parts = []
    parts.append(_("voice_welcome"))
    parts.append(_("voice_ledger"))
    parts.append(_("voice_entries").format(count=len(entries)))
    parts.append(_("voice_credit").format(credit_htg=total_credit_htg, credit_usd=total_credit_usd))
    parts.append(_("voice_expenses").format(expense_htg=total_expense_htg, expense_usd=total_expense_usd))
    parts.append(_("voice_balance").format(balance_htg=net_htg, balance_usd=net_usd))
    parts.append(_("voice_how_it_works"))
    parts.append(_("voice_closing"))
    
    text = " ".join(parts)
    return text

def text_to_speech(text, lang='en'):
    if not EDGE_TTS_AVAILABLE:
        raise RuntimeError("edge-tts not installed.")
    voice_map = {
        'en': 'en-US-JennyNeural',
        'fr': 'fr-FR-DeniseNeural',
        'es': 'es-ES-ElviraNeural'
    }
    voice = voice_map.get(lang, 'en-US-JennyNeural')
    try:
        communicate = edge_tts.Communicate(text, voice)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.mp3') as tmp:
            tmp_path = tmp.name
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(communicate.save(tmp_path))
        loop.close()
        with open(tmp_path, 'rb') as f:
            audio_bytes = f.read()
        os.unlink(tmp_path)
        return audio_bytes
    except Exception as e:
        raise RuntimeError(f"edge-tts generation failed: {e}")

def play_voice_explanation():
    df_rec = get_reconciliation_entries()
    if not df_rec.empty:
        last_row = df_rec.iloc[-1]
        net_usd = last_row['net_usd']
        net_htg = last_row['net_htg']
    else:
        net_usd = 0
        net_htg = 0
    explanation = generate_voice_explanation(df_rec, net_htg, net_usd, selected_lang)
    try:
        audio_bytes = text_to_speech(explanation, lang=selected_lang)
        if audio_bytes:
            st.audio(audio_bytes, format='audio/mp3')
            st.success("Voice explanation played with native voice!")
    except Exception as e:
        st.error(f"Voice generation failed: {e}")

# ----------------------------------------------------------------------
# Excel export
# ----------------------------------------------------------------------
def export_styled_excel(df, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Reconciliation", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Reconciliation"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        currency_fmt = numbers.FORMAT_CURRENCY_USD_SIMPLE
        htg_fmt = '#,##0.00 "G"'
        even_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Header
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in row:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Format specific columns
        headers = [cell.value for cell in worksheet[1]]
        for col_idx, header in enumerate(headers, start=1):
            col_letter = worksheet.cell(row=1, column=col_idx).column_letter
            if header in ['unit usd', 'total usd', 'Credit (Cash In USD)', 'net_usd']:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = currency_fmt
            elif header in ['unit htg', 'total htg', 'Credit (Cash In HTG)', 'net_htg']:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = htg_fmt
        
        # Auto-width
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[col_letter].width = min(adjusted_width, 30)
    
    output.seek(0)
    return output

# ----------------------------------------------------------------------
# Main UI
# ----------------------------------------------------------------------
if not check_password():
    st.stop()

# Language selector
lang_options = {"en": "🇺🇸 English", "fr": "🇫🇷 Français", "es": "🇪🇸 Español"}
if "language" not in st.session_state:
    st.session_state.language = "en"
selected_lang = st.sidebar.selectbox("🌐 Language", options=list(lang_options.keys()),
                                     format_func=lambda x: lang_options[x],
                                     index=["en","fr","es"].index(st.session_state.language))
if selected_lang != st.session_state.language:
    st.session_state.language = selected_lang
    st.rerun()

with st.sidebar:
    st.image(
        "https://raw.githubusercontent.com/Deslandes1/Accountant-Excel-Advanced-AI-/main/Gemini_Generated_Image_8s108y8s108y8s10.png",
        width=100
    )
    st.title(_("app_title"))
    st.markdown("**GlobalInternet.py**")
    st.markdown("Gesner Deslandes")
    st.markdown("Chief Engineer at GlobalInternet.py")
    st.markdown("📧 deslandes78@gmail.com | 📞 (509) 4738-5663")
    st.markdown("---")
    
    if EDGE_TTS_AVAILABLE:
        st.success("✅ Native voice engine (edge-tts) is active")
        auto_speak = st.checkbox("🔊 Auto-speak after update", value=st.session_state.get("auto_speak", False))
        st.session_state.auto_speak = auto_speak
        
        if st.button("🎙️ " + _("Explain Ledger (AI Voice)")):
            play_voice_explanation()
    else:
        st.error("❌ Native voice engine (edge-tts) is NOT installed. Please add 'edge-tts>=6.1.9' to your requirements.txt and redeploy.")
        st.button("🎙️ " + _("Explain Ledger (AI Voice)"), disabled=True)
    
    st.markdown("---")
    if st.button(_("logout")):
        logout()
    st.markdown("---")
    st.markdown("© 2026 GlobalInternet.py – All rights reserved")

# Main header
col1, col2, col3 = st.columns([1, 2, 1])
with col1:
    st.image(
        "https://raw.githubusercontent.com/Deslandes1/Accountant-Excel-Advanced-AI-/main/Gemini_Generated_Image_8s108y8s108y8s10.png",
        width=100
    )
with col2:
    st.markdown(f"<h1 style='text-align: center;'>{_('app_title')}</h1>", unsafe_allow_html=True)
    st.markdown(f"<p style='text-align: center;'><em>{_('subtitle')}</em></p>", unsafe_allow_html=True)
with col3:
    st.markdown("""
    <div style='text-align: right;'>
        <b>GlobalInternet.py</b><br>
        Gesner Deslandes<br>
        Chief Engineer at GlobalInternet.py
    </div>
    """, unsafe_allow_html=True)
st.divider()

# Tabs
tab1, tab2, tab3, tab4, tab5 = st.tabs([_("dashboard"), _("cash_tab"), _("loans_tab"), _("reports_tab"), _("reconciliation_tab")])

# ---- Dashboard ----
with tab1:
    st.header(_("dashboard"))
    balance_usd = get_cash_balance()
    balance_htg = usd_to_htg(balance_usd)
    col1, col2 = st.columns(2)
    with col1:
        st.metric(_("current_balance"), f"${balance_usd:,.2f}")
    with col2:
        st.metric(_("current_balance_htg"), f"G {balance_htg:,.2f}")
    
    col1, col2 = st.columns(2)
    with col1:
        st.subheader(_("recent_transactions"))
        conn = sqlite3.connect("accounting.db")
        recent_cash = pd.read_sql_query("SELECT date, type, category, description, amount FROM cash_transactions ORDER BY date DESC LIMIT 10", conn)
        conn.close()
        if not recent_cash.empty:
            recent_cash['amount_htg'] = recent_cash['amount'].apply(usd_to_htg)
            st.dataframe(recent_cash, use_container_width=True)
        else:
            st.info(_("no_data"))
    with col2:
        st.subheader(_("active_loans"))
        active_loans = get_loans(status='active')
        if not active_loans.empty:
            active_loans['amount_htg'] = active_loans['amount'].apply(usd_to_htg)
            st.dataframe(active_loans[['borrower', 'amount', 'amount_htg', 'payments_made', 'total_payments', 'status']], use_container_width=True)
        else:
            st.info(_("no_active_loans"))

# ---- Cash In/Out ----
with tab2:
    st.header(_("cash_tab"))
    with st.form("cash_form"):
        date = st.date_input(_("date"), value=datetime.date.today())
        trans_type = st.selectbox(_("type"), [_("income"), _("expense")])
        category = st.text_input(_("category"))
        description = st.text_area(_("description"))
        amount = st.number_input(_("amount"), min_value=0.01, step=0.01)
        submitted = st.form_submit_button(_("add_transaction"))
        if submitted:
            add_cash_transaction(str(date), trans_type, category, description, amount)
            st.success(_("transaction_added"))
            st.rerun()
    
    st.subheader(_("transaction_history"))
    conn = sqlite3.connect("accounting.db")
    cash_df = pd.read_sql_query("SELECT * FROM cash_transactions ORDER BY date DESC", conn)
    conn.close()
    if not cash_df.empty:
        cash_df['amount_htg'] = cash_df['amount'].apply(usd_to_htg)
        st.dataframe(cash_df, use_container_width=True)
    else:
        st.info(_("no_data"))
    
    if not cash_df.empty:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            cash_df.to_excel(writer, sheet_name="Cash Transactions", index=False)
        st.download_button(_("download_excel"), data=output.getvalue(), file_name="cash_transactions.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---- Loans ----
with tab3:
    st.header(_("loan_management"))
    with st.expander(_("add_new_loan")):
        with st.form("loan_form"):
            borrower = st.text_input(_("borrower_name"))
            amount = st.number_input(_("loan_amount"), min_value=0.01, step=0.01)
            start_date = st.date_input(_("start_date"), value=datetime.date.today())
            interest_rate = st.number_input(_("interest_rate"), min_value=0.0, step=0.1, value=0.0)
            payment_frequency = st.selectbox(_("payment_frequency"), [_("weekly"), _("monthly")])
            payment_amount = st.number_input(_("payment_amount"), min_value=0.01, step=0.01)
            total_payments = st.number_input(_("total_payments"), min_value=1, step=1, value=12)
            submitted = st.form_submit_button(_("create_loan"))
            if submitted:
                add_loan(borrower, amount, str(start_date), interest_rate, payment_frequency, payment_amount, total_payments)
                st.success(_("loan_created"))
                st.rerun()
    
    st.subheader(_("all_loans"))
    loans_df = get_loans()
    if not loans_df.empty:
        loans_df['amount_htg'] = loans_df['amount'].apply(usd_to_htg)
        loans_df['payment_amount_htg'] = loans_df['payment_amount'].apply(usd_to_htg)
        st.dataframe(loans_df[['id', 'borrower', 'amount', 'amount_htg', 'start_date', 'payment_frequency',
                               'payment_amount', 'payment_amount_htg', 'payments_made', 'total_payments', 'status']],
                     use_container_width=True)
        loan_id = st.selectbox(_("select_loan"), loans_df['id'].tolist())
        loan_data = loans_df[loans_df['id'] == loan_id].iloc[0]
        st.write(f"**{_('borrower_name')}:** {loan_data['borrower']}")
        st.write(f"**{_('remaining_payments')}:** {loan_data['total_payments'] - loan_data['payments_made']}")
        st.write(f"**{_('status')}:** {loan_data['status']}")
        
        if loan_data['status'] == 'active':
            with st.form("payment_form"):
                payment_date = st.date_input(_("payment_date"), value=datetime.date.today())
                payment_amount = st.number_input(_("payment_amount"), value=float(loan_data['payment_amount']), step=0.01)
                if st.form_submit_button(_("record_payment")):
                    record_loan_payment(loan_id, str(payment_date), payment_amount)
                    st.success(_("payment_recorded"))
                    st.rerun()
        
        payments_df = get_loan_payments(loan_id)
        if not payments_df.empty:
            st.subheader(_("payment_history"))
            payments_df['amount_htg'] = payments_df['amount'].apply(usd_to_htg)
            st.dataframe(payments_df, use_container_width=True)
    else:
        st.info(_("no_loans"))

# ---- Reports ----
with tab4:
    st.header(_("generate_reports"))
    report_type = st.selectbox(_("report_type"), [_("cash_flow_statement"), _("loan_status_report"), _("payment_history_report")])
    
    if report_type == _("cash_flow_statement"):
        start_date = st.date_input(_("from_date"), value=datetime.date.today() - datetime.timedelta(days=30))
        end_date = st.date_input(_("to_date"), value=datetime.date.today())
        if st.button(_("generate")):
            df = get_cash_flow(str(start_date), str(end_date))
            st.subheader(f"{_('cash_flow_statement')} {start_date} → {end_date}")
            if not df.empty:
                df['amount_htg'] = df['amount'].apply(usd_to_htg)
                st.dataframe(df, use_container_width=True)
                total_income = df[df['type'] == 'Income']['amount'].sum()
                total_expense = df[df['type'] == 'Expense']['amount'].sum()
                col1, col2, col3 = st.columns(3)
                col1.metric(_("total_income"), f"${total_income:,.2f}")
                col2.metric(_("total_expense"), f"${total_expense:,.2f}")
                col3.metric(_("net_cash_flow"), f"${total_income - total_expense:,.2f}")
                st.metric(_("total_income") + " (HTG)", f"G {usd_to_htg(total_income):,.2f}")
                st.metric(_("total_expense") + " (HTG)", f"G {usd_to_htg(total_expense):,.2f}")
                output_excel = io.BytesIO()
                with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="Cash Flow", index=False)
                st.download_button(_("download_excel"), data=output_excel.getvalue(),
                                   file_name=f"cash_flow_{start_date}_to_{end_date}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                pdf_buffer = generate_pdf_report(f"{_('cash_flow_statement')} {start_date} → {end_date}", df, list(df.columns))
                st.download_button("📄 Download PDF", data=pdf_buffer, file_name=f"cash_flow_{start_date}_to_{end_date}.pdf",
                                   mime="application/pdf")
            else:
                st.info(_("no_data"))
    
    elif report_type == _("loan_status_report"):
        status_filter = st.selectbox(_("filter_by_status"), [_("all"), _("active"), _("completed")])
        if status_filter == _("all"):
            df = get_loans()
        elif status_filter == _("active"):
            df = get_loans(status='active')
        else:
            df = get_loans(status='completed')
        if st.button(_("generate")):
            if not df.empty:
                df['amount_htg'] = df['amount'].apply(usd_to_htg)
                st.dataframe(df, use_container_width=True)
                output_excel = io.BytesIO()
                with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name="Loans", index=False)
                st.download_button(_("download_excel"), data=output_excel.getvalue(), file_name="loan_report.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                pdf_buffer = generate_pdf_report(_("loan_status_report"), df, list(df.columns))
                st.download_button("📄 Download PDF", data=pdf_buffer, file_name="loan_report.pdf", mime="application/pdf")
            else:
                st.info(_("no_data"))
    
    else:  # payment history report
        all_loans = get_loans()
        if not all_loans.empty:
            selected_loan = st.selectbox(_("select_loan_for_history"), all_loans['id'].tolist(),
                                         format_func=lambda x: f"Loan #{x} - {all_loans[all_loans['id']==x]['borrower'].values[0]}")
            if st.button(_("generate")):
                payments = get_loan_payments(selected_loan)
                if not payments.empty:
                    payments['amount_htg'] = payments['amount'].apply(usd_to_htg)
                    st.dataframe(payments, use_container_width=True)
                    output_excel = io.BytesIO()
                    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                        payments.to_excel(writer, sheet_name="Payments", index=False)
                    st.download_button(_("download_excel"), data=output_excel.getvalue(),
                                       file_name=f"loan_{selected_loan}_payments.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    pdf_buffer = generate_pdf_report(f"{_('payment_history_report')} Loan #{selected_loan}", payments,
                                                     list(payments.columns))
                    st.download_button("📄 Download PDF", data=pdf_buffer,
                                       file_name=f"loan_{selected_loan}_payments.pdf", mime="application/pdf")
                else:
                    st.info(_("no_data"))
        else:
            st.info(_("no_loans"))

# ---- Reconciliation Ledger (HARDCODED ENGLISH – NO TRANSLATION KEYS) ----
with tab5:
    st.header("📋 Reconciliation Ledger")
    st.caption("Exchange Rate: 1 USD = 100 HTG")
    st.info("💡 How it works: Enter Credit (Cash In) in HTG. The system converts it to USD at 1 USD = 100 HTG. Expenses reduce the net balance.")

    # ---- RESET BUTTON ----
    col_reset, _ = st.columns([1, 3])
    with col_reset:
        if st.button("🗑️ Reset Ledger (Clear All Entries)", use_container_width=True):
            if st.checkbox("⚠️ Confirm: delete ALL entries?"):
                reset_ledger()
                st.success("Ledger reset! Start adding your own entries.")
                st.rerun()
            else:
                st.warning("Please confirm the deletion.")

    df_rec = get_reconciliation_entries()
    
    if not df_rec.empty:
        last_row = df_rec.iloc[-1]
        net_usd = last_row['net_usd']
        net_htg = last_row['net_htg']
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Net Balance (USD)", f"${net_usd:,.2f}")
        with col2:
            st.metric("Net Balance (HTG)", f"G {net_htg:,.2f}")

        # Summary
        st.subheader("📊 Cash In / Expenses Summary")
        credit_total_htg = df_rec['credit'].sum()
        expense_total_htg = df_rec['total_htg'].sum()
        net_htg_summary = credit_total_htg - expense_total_htg

        credit_total_usd = credit_total_htg / EXCHANGE_RATE
        expense_total_usd = expense_total_htg / EXCHANGE_RATE
        net_usd_summary = credit_total_usd - expense_total_usd

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("💰 Total Cash In (HTG)", f"G {credit_total_htg:,.2f}")
            st.metric("💰 Total Cash In (USD)", f"${credit_total_usd:,.2f}")
        with col2:
            st.metric("💸 Total Expenses (HTG)", f"G {expense_total_htg:,.2f}")
            st.metric("💸 Total Expenses (USD)", f"${expense_total_usd:,.2f}")
        with col3:
            st.metric("📊 Net Balance (HTG)", f"G {net_htg_summary:,.2f}", delta=f"{net_htg_summary:,.2f}")
            st.metric("📊 Net Balance (USD)", f"${net_usd_summary:,.2f}", delta=f"{net_usd_summary:,.2f}")

        # Manual calculator
        st.markdown("---")
        st.subheader("🧮 Quick Cash Calculator (Manual Entry)")
        st.caption("Enter any amounts below to calculate Cash In - Expenses. This does NOT affect your ledger.")

        col_curr, col_in, col_exp = st.columns([1, 2, 2])
        with col_curr:
            calc_currency = st.selectbox("Currency", ["HTG (G)", "USD ($)"], key="calc_currency")
        with col_in:
            calc_cashin = st.number_input("💰 Total Cash In", min_value=0.0, step=100.0, value=0.0, key="calc_cashin")
        with col_exp:
            calc_expenses = st.number_input("💸 Total Expenses", min_value=0.0, step=100.0, value=0.0, key="calc_expenses")

        if st.button("🧮 Calculate Net", key="calc_btn"):
            net = calc_cashin - calc_expenses
            currency_symbol = "G" if calc_currency == "HTG (G)" else "$"
            st.markdown("---")
            col_res1, col_res2 = st.columns(2)
            with col_res1:
                st.metric("Cash In", f"{currency_symbol} {calc_cashin:,.2f}")
                st.metric("Expenses", f"{currency_symbol} {calc_expenses:,.2f}")
            with col_res2:
                st.metric("Net Balance (Result)", f"{currency_symbol} {net:,.2f}", 
                          delta=f"{net:,.2f}", delta_color="normal" if net >= 0 else "inverse")
    else:
        st.info("No data available.")
    
    st.subheader("📋 Reconciliation Table")
    
    if not df_rec.empty:
        display_cols = ['id', 'date', 'credit', 'description', 'qty', 'unit_htg', 'unit_usd', 
                        'total_htg', 'total_usd', 'net_htg', 'net_usd']
        df_display = df_rec[display_cols].copy()
        
        col_headers = {
            'id': 'ID',
            'date': 'Date',
            'credit': 'Credit (Cash In HTG)',
            'description': 'Description / Item Details',
            'qty': 'Qty',
            'unit_htg': 'Unit Price (HTG)',
            'unit_usd': 'Unit Price (USD)',
            'total_htg': 'Total (HTG)',
            'total_usd': 'Total (USD)',
            'net_htg': 'Net Balance (HTG)',
            'net_usd': 'Net Balance (USD)'
        }
        df_display.rename(columns=col_headers, inplace=True)
        
        column_config = {
            'ID': st.column_config.NumberColumn('ID', format="%d"),
            'Date': st.column_config.TextColumn('Date'),
            'Credit (Cash In HTG)': st.column_config.NumberColumn('Credit (Cash In HTG)', format="G %,.2f"),
            'Description / Item Details': st.column_config.TextColumn('Description / Item Details'),
            'Qty': st.column_config.NumberColumn('Qty', format="%f"),
            'Unit Price (HTG)': st.column_config.NumberColumn('Unit Price (HTG)', format="G %,.2f"),
            'Unit Price (USD)': st.column_config.NumberColumn('Unit Price (USD)', format="$% ,.2f"),
            'Total (HTG)': st.column_config.NumberColumn('Total (HTG)', format="G %,.2f"),
            'Total (USD)': st.column_config.NumberColumn('Total (USD)', format="$% ,.2f"),
            'Net Balance (HTG)': st.column_config.NumberColumn('Net Balance (HTG)', format="G %,.2f"),
            'Net Balance (USD)': st.column_config.NumberColumn('Net Balance (USD)', format="$% ,.2f")
        }
        
        st.dataframe(df_display, column_config=column_config, use_container_width=True, hide_index=True)
    else:
        st.info("No data available.")
    
    # ---- ADD ENTRY SECTION WITH FORMULA REMINDER ----
    col_title, col_formula = st.columns([1, 2])
    with col_title:
        st.subheader("➕ Add Entry")
    with col_formula:
        st.markdown(
            """
            <div style='text-align: right; margin-top: 15px;'>
                <span style='background-color: #d9eaf7; padding: 4px 14px; border-radius: 20px; font-weight: 500; color: #003366; border: 1px solid #99ccff; font-size: 14px;'>
                    📐 (Net Balance = Sum of all Cash In – Sum of all Expenses)
                </span>
            </div>
            """,
            unsafe_allow_html=True
        )
    # -----------------------------------------------------
    
    with st.form("reconciliation_form"):
        col1, col2 = st.columns(2)
        with col1:
            date = st.date_input("Date", value=datetime.date.today())
            credit_htg = st.number_input("Credit (Cash In HTG)", min_value=0.0, step=0.01, value=0.0)
            description = st.text_input("Description / Item Details")
        with col2:
            qty = st.number_input("Qty", min_value=0.0, step=0.01, value=0.0, key="qty_input")
            unit_htg = st.number_input("Unit Price (HTG)", min_value=0.0, step=0.01, value=0.0, key="unit_htg_input")
        
        qty_val = st.session_state.get("qty_input", 0.0)
        unit_htg_val = st.session_state.get("unit_htg_input", 0.0)
        credit_usd = credit_htg / EXCHANGE_RATE
        unit_usd_preview = unit_htg_val / EXCHANGE_RATE
        total_htg_preview = qty_val * unit_htg_val
        total_usd_preview = qty_val * unit_usd_preview
        
        st.markdown("---")
        st.markdown("**📊 Preview (will be used when you submit)**")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Credit (USD)", f"${credit_usd:.2f}")
        col2.metric("Unit Price (USD)", f"{unit_usd_preview:.2f}")
        col3.metric("Total (HTG)", f"{total_htg_preview:.2f}")
        col4.metric("Total (USD)", f"{total_usd_preview:.2f}")
        
        submitted = st.form_submit_button("Add Entry")
        if submitted:
            if description.strip() == "":
                st.error("Description is required.")
            else:
                add_reconciliation_entry(str(date), credit_htg, description, qty_val, unit_htg_val, unit_usd_preview, total_htg_preview, total_usd_preview)
                st.success("Entry added!")
                if st.session_state.get("auto_speak", False):
                    play_voice_explanation()
                st.rerun()
    
    if not df_rec.empty:
        st.subheader("🗑️ Delete Entry")
        delete_id = st.selectbox("Select entry ID to delete", df_rec['id'].tolist(),
                                 format_func=lambda x: f"ID {x} - {df_rec[df_rec['id']==x]['description'].iloc[0]}")
        if st.button("Delete selected entry", use_container_width=True):
            if delete_id == 1:
                st.error("Cannot delete the initial balance row.")
            else:
                delete_reconciliation_entry(delete_id)
                st.success("Entry deleted.")
                if st.session_state.get("auto_speak", False):
                    play_voice_explanation()
                st.rerun()
    
    if not df_rec.empty:
        styled_excel = export_styled_excel(df_rec, "Reconciliation July - 2026")
        st.download_button(
            "📥 Download Excel",
            data=styled_excel,
            file_name="reconciliation_ledger.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
